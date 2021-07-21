from datetime import datetime, date
from flask import Blueprint, url_for, render_template, request, flash
from sqlalchemy.sql import func
from sqlalchemy import literal

from lease import db
from lease.models import Contract, Baseline, Schedule, Leasecls
from lease.forms import Report
from ifrs16 import date_fromto_compare

import sqlite3

import os
import getpass
from openpyxl import Workbook


bp = Blueprint('report', __name__, url_prefix='/report')


@bp.route('/sum/', methods=('GET', 'POST'))
def sum():
    report = Report()
    if request.method == 'POST':
        report.frdate = request.values.get("frdate")
        report.todate = request.values.get("todate")

        if report.frdate != "" and report.todate != "":
            if report.frdate != report.todate:
                date_compare = date_fromto_compare(date.fromisoformat(report.frdate), date.fromisoformat(report.todate))
                if date_compare == False:
                    flash('조회기간 종료일은 시작일보다 커야 합니다.')
                    return render_template('contract/month_report.html',report=report)

        # 조회기간 비용수익 조회
        list_sql = db.session.query(Contract.leasecls_id, Leasecls.clsname,
                                    func.sum(Schedule.lease_fee).label("lease_fee"),
                                    func.sum(Schedule.lease_exp).label("lease_exp"),
                                    func.sum(Schedule.deposit_intr).label("deposit_intr"),
                                    func.sum(Schedule.lease_dep).label("lease_dep"),
                                    func.sum(Schedule.rstr_exp).label("rstr_exp")).\
            join(Contract, Schedule.contract_id == Contract.id).\
            outerjoin(Leasecls, Leasecls.id == Contract.leasecls_id).\
            filter(Schedule.exemonth.between(report.frdate, report.todate)).\
            group_by(Contract.leasecls_id)
        sum_sql = db.session.query(literal("0").label("leasecls_id"),literal("합계").label("clsname"),
                                   func.sum(Schedule.lease_fee).label("lease_fee"),
                                   func.sum(Schedule.lease_exp).label("lease_exp"),
                                   func.sum(Schedule.deposit_intr).label("deposit_intr"),
                                   func.sum(Schedule.lease_dep).label("lease_dep"),
                                   func.sum(Schedule.rstr_exp).label("rstr_exp")).\
            join(Contract, Schedule.contract_id == Contract.id).\
            filter(Schedule.exemonth.between(report.frdate,report.todate))
        prdpl = list_sql.union_all(sum_sql).all()


        # 잔액조회
        list_sql = db.session.query(Contract.leasecls_id,Leasecls.clsname,
                                    func.sum(Schedule.lease_lbt).label("lease_lbt"),
                                    func.sum(Schedule.deposit_dscnt).label("deposit_dscnt"),
                                    func.sum(Schedule.rstr_alw).label("rstr_alw"),
                                    func.sum(Schedule.lease_acq).label("lease_acq"),
                                    func.sum(Schedule.dep_accum).label("dep_accum")).\
            join(Contract, Schedule.contract_id == Contract.id).\
            outerjoin(Leasecls, Leasecls.id == Contract.leasecls_id).\
            filter(Schedule.exemonth == report.todate).\
            group_by(Contract.leasecls_id)
        sum_sql = db.session.query(literal("0").label("leasecls_id"),literal("합계").label("clsname"),
                                   func.sum(Schedule.lease_lbt).label("lease_lbt"),
                                   func.sum(Schedule.deposit_dscnt).label("deposit_dscnt"),
                                   func.sum(Schedule.rstr_alw).label("rstr_alw"),
                                   func.sum(Schedule.lease_acq).label("lease_acq"),
                                   func.sum(Schedule.dep_accum).label("dep_accum")).\
            join(Contract, Schedule.contract_id == Contract.id).\
            filter(Schedule.exemonth == report.todate)
        balance = list_sql.union_all(sum_sql).all()

        conn = sqlite3.connect("lease.db")
        cur = conn.cursor()

        # 마이그레이션 후, 계산 집계
        cur.execute(
            "SELECT sum(lease_lbt), sum(deposit_dscnt), sum(rstr_alw), sum(lease_acq), sum(dep_accum) "
            "FROM schedule WHERE cycle = '0' AND exemonth = '2021-04-01';")
        result = cur.fetchone()
        report.mig_lease_lbt = result[0]
        report.mig_deposit_dscnt = result[1]
        report.mig_rstr_alw = result[2]
        report.mig_lease_ast = result[3]
        report.mig_dep_accum = result[4]

        # 마이그레이션 값 집계
        cur.execute(
            "SELECT sum(prnt_lblti),sum(prnt_dpsit),sum(prnt_rstra),(sum(prnt_asset)+sum(prnt_depac)) as sumv,sum(prnt_depac) "
            "FROM baseline WHERE startdate = '2021-04-01';")
        result = cur.fetchone()
        report.prnt_lblti = result[0]
        report.prnt_dpsit = result[1]
        report.prnt_rstra = result[2]
        report.prnt_asset = result[3]
        report.prnt_depac = result[4]

        conn.close()

        return render_template('contract/report_sum.html',report=report, balance=balance, prdpl=prdpl)

    return render_template('contract/report_sum.html',report=report)


@bp.route('/items/', methods=('GET', 'POST'))
def items():
    report = Report()
    if request.method == 'POST':
        report.msgtype = ""
        report.todate = request.values.get("todate")

        # 잔액조회
        list_sql = db.session.query(Contract.leasecls_id,Leasecls.clsname,Contract.id,Contract.subject,
                                    Schedule.lease_fee,Schedule.lease_exp,Schedule.lease_lbt,
                                    Schedule.deposit_dscnt,Schedule.deposit_intr,
                                    Schedule.rstr_exp, Schedule.rstr_alw,
                                    Schedule.lease_dep, Schedule.dep_accum, Schedule.lease_acq).\
            join(Baseline, Schedule.contract_id == Baseline.contract_id).\
            join(Contract, Schedule.contract_id == Contract.id).\
            outerjoin(Leasecls, Leasecls.id == Contract.leasecls_id).\
            filter(Schedule.exemonth == report.todate,
                   Baseline.startdate <= report.todate, Baseline.enddate >= report.todate)
        sum_sql = db.session.query(literal("0").label("leasecls_id"),literal("구분").label("clsname"),
                                   literal("0").label("id"),literal("합계").label("subject"),
                                   func.sum(Schedule.lease_fee).label("lease_fee"),
                                   func.sum(Schedule.lease_exp).label("lease_exp"),
                                   func.sum(Schedule.lease_lbt).label("lease_lbt"),
                                   func.sum(Schedule.deposit_dscnt).label("deposit_dscnt"),
                                   func.sum(Schedule.deposit_intr).label("deposit_intr"),
                                   func.sum(Schedule.rstr_exp).label("rstr_exp"),
                                   func.sum(Schedule.rstr_alw).label("rstr_alw"),
                                   func.sum(Schedule.lease_dep).label("lease_dep"),
                                   func.sum(Schedule.dep_accum).label("dep_accum"),
                                   func.sum(Schedule.lease_acq).label("lease_acq")).\
            join(Baseline, Schedule.contract_id == Baseline.contract_id).\
            join(Contract, Schedule.contract_id == Contract.id).\
            filter(Schedule.exemonth == report.todate,
                   Baseline.startdate <= report.todate, Baseline.enddate >= report.todate)
        balance = sum_sql.union_all(list_sql).all()

        postid = request.form.get("download", "")
        if postid == "다운로드":
            if len(balance) == 0:
                flash("다운로드할 데이터가 없습니다.")
                return render_template('contract/report_items.html',report=report)

            # 바탕화면에 다운로드 폴더 생성
            username = getpass.getuser()
    
            try:
                dir_path = os.path.join("C:\\Users", username, "Desktop", "다운로드")
                if not (os.path.isdir(dir_path)):
                    os.makedirs((os.path.join(dir_path)))
            except:
                flash("바탕화면에 폴더생성 실패")
                return render_template('contract/report_items.html',report=report)

            # 엑셀파일 생성
            write_wb = Workbook()

            # 이름이 있는 시트를 생성
            # write_ws = write_wb.create_sheet('생성시트')
            write_ws = write_wb.active
            write_ws.freeze_panes = 'B2'

            # 엑셀 값 입력
            for row_idx, val_line in enumerate(balance):
                if row_idx == 0:
                    # 행 단위로 추가
                    xlsheader = ["구분","구분명","번호","제목","리스료","이자비용","리스부채","보증금현할차",
                                 "보증금유효이자","충당부채전입액","복구충당부채","자산상각비","상각누계액","사용권자산"]
                    write_ws.append(xlsheader)
                else:
                    # 셀 단위로 추가
                    for col_idx, cell in enumerate(val_line):
                        row = row_idx + 1
                        col = col_idx + 1
                        if col > 4:
                            write_ws.cell(row, col, cell).number_format = '#,##0'
                        else:
                            write_ws.cell(row, col, cell)

            file_name = str(report.todate) + ".xlsx"
            # file_path = os.path.join(dir_path, "다운로드.xlsx")
            file_path = os.path.join(dir_path, file_name)


            # 파일 저장
            # write_wb.save(file_path)
            try:
                write_wb.save(file_path)
                flash("바탕화면/다운로드 폴더에 파일이 생성되었습니다.")
                report.msgtype = "success"
            except:
                flash("파일 생성 실패")

        return render_template('contract/report_items.html',report=report, balance=balance)

    return render_template('contract/report_items.html',report=report)


@bp.route('/report2/', methods=('GET', 'POST'))
def report2():
    report = Report()
    if request.method == 'POST':
        report.frdate = request.values.get("frdate")
        report.todate = request.values.get("todate")

        if report.frdate != "" and report.todate != "":
            if report.frdate != report.todate:
                date_compare = date_fromto_compare(date.fromisoformat(report.frdate), date.fromisoformat(report.todate))
                if date_compare == False:
                    flash('조회기간 종료일은 시작일보다 커야 합니다.')
                    return render_template('contract/month_report.html',report=report)

        conn = sqlite3.connect("lease.db")
        cur = conn.cursor()

        param_range = []
        param_range.append(str(report.frdate))
        param_range.append(str(report.todate))

        param = []
        param.append(str(report.todate))


        # param = ['2021-04-01', '2021-06-30']

        # 월 상각비 집계
        cur.execute(
            "SELECT sum(lease_exp),sum(deposit_intr),sum(lease_dep) "
            "FROM schedule WHERE exemonth BETWEEN ? AND ?;",param_range)
        result = cur.fetchone()
        report.lease_exp = result[0]
        report.deposit_intr = result[1]
        report.lease_dep = result[2]

        # 월 시점 잔액
        cur.execute(
            "SELECT sum(lease_lbt), sum(deposit_dscnt), sum(rstr_alw), sum(lease_acq), sum(dep_accum) "
            "FROM schedule WHERE exemonth = ?;",param)
        result = cur.fetchone()
        report.lease_lbt = result[0]
        report.deposit_dscnt = result[1]
        report.rstr_alw = result[2]
        report.lease_ast = result[3]
        report.dep_accum = result[4]

        # 마이그레이션 후, 계산 집계
        cur.execute(
            "SELECT sum(lease_lbt), sum(deposit_dscnt), sum(rstr_alw), sum(lease_acq), sum(dep_accum) "
            "FROM schedule WHERE cycle = '0' AND exemonth BETWEEN ? AND ?;",param_range)
        result = cur.fetchone()
        report.mig_lease_lbt = result[0]
        report.mig_deposit_dscnt = result[1]
        report.mig_rstr_alw = result[2]
        report.mig_lease_ast = result[3]
        report.mig_dep_accum = result[4]

        # 마이그레이션 값 집계
        cur.execute(
            "SELECT sum(prnt_lblti),sum(prnt_dpsit),sum(prnt_rstra),(sum(prnt_asset)+sum(prnt_depac)) as sumv,sum(prnt_depac) "
            "FROM baseline WHERE startdate BETWEEN ? AND ?;",param_range)
        result = cur.fetchone()
        report.prnt_lblti = result[0]
        report.prnt_dpsit = result[1]
        report.prnt_rstra = result[2]
        report.prnt_asset = result[3]
        report.prnt_depac = result[4]

        conn.close()

        return render_template('contract/month_report.html',report=report)

    return render_template('contract/month_report.html',report=report)